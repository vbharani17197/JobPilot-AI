"""Excel workbook generation (the primary deliverable).

Four worksheets, exactly per spec:
  1. Top 20 Jobs Overall
  2. Top 20 Direct Company Career Site Jobs
  3. New Jobs Since Previous Run
  4. Skill Insights & ATS Recommendations  (ATS recs ONLY here)
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .config import Config
from .insights import Insights
from .logging_setup import logger
from .models import Job

# ---- styling constants ----------------------------------------------
_FONT = "Calibri"
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(name=_FONT, bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(name=_FONT, bold=True, size=14, color="1F3864")
_SUBHEAD_FONT = Font(name=_FONT, bold=True, size=12, color="1F3864")
_CELL_FONT = Font(name=_FONT, size=10)
_LINK_FONT = Font(name=_FONT, size=10, color="0563C1", underline="single")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALT_FILL = PatternFill("solid", fgColor="F2F5FB")

_SENTIMENT_FILL = {
    "Positive": PatternFill("solid", fgColor="C6EFCE"),
    "Negative": PatternFill("solid", fgColor="FFC7CE"),
    "Neutral": PatternFill("solid", fgColor="FFEB9C"),
}


def generate_report(jobs: list[Job], new_jobs: list[Job], insights: Insights,
                    config: Config) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    ranked = sorted(jobs, key=lambda j: j.match_score, reverse=True)

    _sheet_overall(wb, ranked, config["output"]["top_overall"])
    _sheet_company_site(wb, ranked, config["output"]["top_company_site"])
    _sheet_new(wb, new_jobs)
    _sheet_insights(wb, insights)

    out_dir = config.output_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = f"Job_Search_Report_{date.today():%Y_%m_%d}.xlsx"
    path = out_dir / fname
    wb.save(path)
    logger.info("Excel report written to {}.", path)
    return path


# ---------------------------------------------------------------------
def _header_row(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[row].height = 28


def _style_data_cell(cell, alt: bool):
    cell.font = _CELL_FONT
    cell.border = _BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if alt:
        cell.fill = _ALT_FILL


def _autosize(ws, widths: dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _hyperlink(cell, url: str):
    if url:
        cell.value = "Apply"
        cell.hyperlink = url
        cell.font = _LINK_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


# ---- Worksheet 1 -----------------------------------------------------
def _sheet_overall(wb, ranked: list[Job], top_n: int):
    ws = wb.create_sheet("Top 20 Overall")
    headers = ["Rank", "Company", "Role", "Location", "Experience Required",
               "Salary", "Match Score", "AI Fit", "Rating", "Sentiment",
               "Apply Type", "Why It Fits (AI)", "Final Apply URL", "Source"]
    _header_row(ws, headers)

    for i, job in enumerate(ranked[:top_n], start=1):
        r = i + 1
        alt = i % 2 == 0
        ai_fit = "" if job.llm_fit_score is None else round(job.llm_fit_score)
        vals = [i, job.company, job.role, job.location, job.experience_required,
                job.salary, job.match_score, ai_fit, job.rating, job.sentiment,
                job.apply_type, job.llm_rationale, None, job.source]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            _style_data_cell(cell, alt)
            if c in (1, 7, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == 10 and job.sentiment in _SENTIMENT_FILL:
                cell.fill = _SENTIMENT_FILL[job.sentiment]
        _hyperlink(ws.cell(row=r, column=13), job.best_url)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{min(len(ranked), top_n) + 1}"
    _autosize(ws, {1: 6, 2: 24, 3: 28, 4: 18, 5: 16, 6: 13, 7: 12, 8: 8,
                   9: 8, 10: 10, 11: 20, 12: 44, 13: 12, 14: 9})


# ---- Worksheet 2 -----------------------------------------------------
def _sheet_company_site(wb, ranked: list[Job], top_n: int):
    ws = wb.create_sheet("Top 20 Company Sites")
    headers = ["Rank", "Company", "Role", "Location", "Experience Required",
               "Salary", "Match Score", "Rating", "Sentiment",
               "Final Apply URL", "Source"]
    _header_row(ws, headers)

    filtered = [j for j in ranked if j.apply_type == "Company Career Site"]
    for i, job in enumerate(filtered[:top_n], start=1):
        r = i + 1
        alt = i % 2 == 0
        vals = [i, job.company, job.role, job.location, job.experience_required,
                job.salary, job.match_score, job.rating, job.sentiment,
                None, job.source]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            _style_data_cell(cell, alt)
            if c in (1, 7):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == 9 and job.sentiment in _SENTIMENT_FILL:
                cell.fill = _SENTIMENT_FILL[job.sentiment]
        _hyperlink(ws.cell(row=r, column=10), job.best_url)

    if not filtered:
        ws.cell(row=2, column=1,
                value="No direct company career-site jobs resolved in this run.")
        ws.merge_cells("A2:K2")

    ws.freeze_panes = "A2"
    _autosize(ws, {1: 6, 2: 26, 3: 30, 4: 20, 5: 18, 6: 14, 7: 12,
                   8: 9, 9: 11, 10: 12, 11: 10})


# ---- Worksheet 3 -----------------------------------------------------
def _sheet_new(wb, new_jobs: list[Job]):
    ws = wb.create_sheet("New Jobs")
    headers = ["Company", "Role", "Date Found", "Final Apply URL"]
    _header_row(ws, headers)

    ranked_new = sorted(new_jobs, key=lambda j: j.match_score, reverse=True)
    for i, job in enumerate(ranked_new, start=1):
        r = i + 1
        alt = i % 2 == 0
        vals = [job.company, job.role, job.first_seen or date.today().isoformat(),
                None]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            _style_data_cell(cell, alt)
        _hyperlink(ws.cell(row=r, column=4), job.best_url)

    if not ranked_new:
        ws.cell(row=2, column=1, value="No new jobs since the previous run.")
        ws.merge_cells("A2:D2")

    ws.freeze_panes = "A2"
    _autosize(ws, {1: 28, 2: 34, 3: 14, 4: 12})


# ---- Worksheet 4 -----------------------------------------------------
def _sheet_insights(wb, ins: Insights):
    ws = wb.create_sheet("Skill Insights & ATS")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 14

    row = 1
    title = ws.cell(row=row, column=2, value="Skill Insights & ATS Recommendations")
    title.font = _TITLE_FONT
    row += 2

    row = _insight_block(ws, row, "Most Requested Skills (you have these)",
                         ins.most_requested_skills, "Mentions")
    row = _insight_block(ws, row, "Trending Technologies",
                         ins.trending_technologies, "Mentions")
    row = _insight_block(ws, row, "Missing Skills (in demand, not on resume)",
                         ins.missing_skills, "Mentions")
    row = _insight_block(ws, row, "Frequently Requested Certifications",
                         ins.certifications, "Mentions")

    # ATS recommendations (text list).
    sub = ws.cell(row=row, column=2, value="ATS Keyword Recommendations")
    sub.font = _SUBHEAD_FONT
    row += 1
    if ins.ats_recommendations:
        for rec in ins.ats_recommendations:
            c = ws.cell(row=row, column=2, value=f"• {rec}")
            c.font = _CELL_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 30
            row += 1
    else:
        ws.cell(row=row, column=2, value="• No recommendations generated this run.")
        row += 1


def _insight_block(ws, row, title, pairs, count_label):
    sub = ws.cell(row=row, column=2, value=title)
    sub.font = _SUBHEAD_FONT
    row += 1
    # mini header
    h1 = ws.cell(row=row, column=2, value="Item")
    h2 = ws.cell(row=row, column=3, value=count_label)
    for h in (h1, h2):
        h.font = _HEADER_FONT
        h.fill = _HEADER_FILL
        h.alignment = Alignment(horizontal="left")
    row += 1
    if not pairs:
        ws.cell(row=row, column=2, value="(none detected)").font = _CELL_FONT
        return row + 2
    for name, cnt in pairs:
        ws.cell(row=row, column=2, value=name).font = _CELL_FONT
        cc = ws.cell(row=row, column=3, value=cnt)
        cc.font = _CELL_FONT
        cc.alignment = Alignment(horizontal="center")
        row += 1
    return row + 1
